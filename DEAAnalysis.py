import pandas as pd
from openpyxl.utils import get_column_letter
from pulp import *
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment


dir_data_prod = "data/prod"
dir_data_test = "data/test"
dir_data_in_use = dir_data_test


file_src = 'DEA.txt'
file_stg = 'DEA_stg.txt'
file_in_use = file_src

file_preprocessed = 'DEA_processed.csv'
file_optimized_weight_factor = 'DEA_weightFactors.csv'
file_optimized_LPModel = 'DEA_LPModel.csv'

file_dea_report = 'DEA_Report.xlsx'

df_input = None
df_output = None

inputName = []
outputName = []
# Load data
def load_data():
    print("--------------Loading data----------------")
    global file_in_use
    lines = None
    lines_stg = []
    with open(f'{dir_data_in_use}/{file_in_use}','r') as file:
        lines  = file.readlines()

    #lines = [i for i in lines if lines !='\n']
    for line in lines:
        if line == '\n':
            pass
        else:
            lines_stg.append(line.strip())
    print(lines_stg)

    file_in_use = file_stg
    with open(f'{dir_data_in_use}/{file_in_use}','w') as file:
        for line in lines_stg:
            file.write(f"{line}\n")
    print('--------------Finished loading data----------------')


# pre-process data
def preprocess_data():
    print("preprocess_data")
    global file_in_use
    global inputName
    global outputName
    global unitName
    file_in_use = file_stg
    unitName = []
    inputUsed = []
    outputProduced = []
    with open(f'{dir_data_in_use}/{file_in_use}','r') as file:
        line = file.readline().strip()
        inputName = line.split(',')
        # print(inputName)
        line = file.readline().strip()
        outputName = line.split(',')
        # print(outputName)

        while file:
            line = file.readline().strip()
            if line == '':
                break
            unitName.append(line)
            line = file.readline().strip()
            line = line.split(',')
            line = [float(i) for i in line]
            inputUsed.append(line)

            line = file.readline().strip()
            line = line.split(',')
            line = [float(i) for i in line]
            outputProduced.append(line)

    print(unitName)
    print(inputUsed)
    print(outputProduced)

    df_input = pd.DataFrame(data= inputUsed, columns = inputName, index = unitName)
    df_output = pd.DataFrame(data=outputProduced, columns=outputName, index=unitName)

    pd.concat([df_input, df_output], axis=1).to_csv(f'{dir_data_in_use}/{file_preprocessed}')
    print(pd.concat([df_input,df_output],axis=1))


# build model
def build_model():
    print("----------build_model----------")
    global file_in_use
    file_in_use = file_preprocessed
    df = pd.read_csv(f'{dir_data_in_use}/{file_in_use}', index_col = 0)
    # print(inputName)

    in_vars = LpVariable.dict('I', inputName, lowBound=0, cat= 'Continuous')
    out_vars = LpVariable.dict('O', outputName, lowBound=0, cat='Continuous')
    df_weightFactor = pd.DataFrame(index = df.index, columns = inputName + outputName)


    for unit in df.index:
        prob = LpProblem("DEA Analysis", LpMaximize)
        for j in range(df.shape[0]):
            prob += (
                lpSum([df.loc[df.index[j],i] * in_vars[i] for i in inputName]) >=
                lpSum([df.loc[df.index[j],i] * out_vars[i] for i in outputName])
                  )

        prob += lpSum([df.loc[unit,i] * in_vars[i] for i in inputName]) == 1
        prob += lpSum([df.loc[unit,i] * out_vars[i] for i in outputName])

        status = prob.solve()
        #print(LpStatus[status])
        weightFactorDict = {}
        for v in prob.variables():
            weightFactorDict[v.name] = v.varValue
        #print(v.name, v.varValue)

        df_weightFactor.loc[unit] = [weightFactorDict['I_Faculty'],
                                 weightFactorDict['I_Support_Staff'],
                                 weightFactorDict['I_Supply_Budget'],
                                 weightFactorDict['O_Credit_Hours'],
                                 weightFactorDict['O_Research_Pubs'],
                                 ]


    df_weightFactor.to_csv(f'{dir_data_in_use}/{file_optimized_weight_factor}')


def prepare_report_data():
    print('-----------generate_report data------------')
    global file_in_use
    file_in_use = file_optimized_weight_factor
    df_weightFactor = pd.read_csv(f'{dir_data_in_use}/{file_in_use}', index_col = 0)
    print(df_weightFactor)

    file_in_use = file_preprocessed
    df_preprocessed = pd.read_csv(f'{dir_data_in_use}/{file_in_use}', index_col=0)
    print(df_preprocessed)


    df_LPModel = df_weightFactor * df_preprocessed
    df_LPModel.to_csv(f'{dir_data_in_use}/{file_optimized_LPModel}')


def generate_report():
    wb = Workbook()
    ws = wb.active

    for row in ws.iter_rows():
        for cell in row:
            cell.value = None

    row_num = 4
    col_num = 2
    ws.title = "DEA Report"
    loc_title_row = row_num - 3
    loc_title_col_char = get_column_letter(col_num)
    ws[f'{loc_title_col_char}{loc_title_row}'] = 'Summary of analysis'
    ws[f'{loc_title_col_char}{loc_title_row}'].font = Font(bold=True)

    build_report_section1(ws, [row_num, col_num])
    build_report_section2(ws, [row_num, col_num + 4])
    build_report_section3(ws, [row_num + 8, col_num + 4])

    wb.save(f'{dir_data_in_use}/{file_dea_report}')


def build_report_section1(ws, loc):
    file_in_use = file_optimized_LPModel
    df_LPModel = pd.read_csv(f'{dir_data_in_use}/{file_in_use}', index_col = 0)

    df_LPModel['LPMax'] = df_LPModel[outputName].sum(axis=1)
    df_LPModel['eff_index'] = df_LPModel['LPMax'].apply(lambda x: 'Yes' if x >= 1 else 'No')
    df_eff = df_LPModel[['LPMax', 'eff_index']]

    row_num, col_num = loc

    loc_eff_row = row_num

    # --- Title row ---
    headers = ['Units', 'LP maximum output', 'Efficient?']
    for i, title in enumerate(headers):
        loc_eff_col_char = get_column_letter(col_num + i)
        ws[f'{loc_eff_col_char}{loc_eff_row}'] = title

        ws[f'{loc_eff_col_char}{loc_eff_row}'].fill = PatternFill(start_color="000080", end_color="000080",
                                                                  fill_type="solid")
        ws[f'{loc_eff_col_char}{loc_eff_row}'].font = Font(color="FFFFFF", bold=True)

    # --- Body rows ---
    for i in range(df_eff.shape[0]):
        values = [df_eff.index[i], df_eff.loc[df_eff.index[i], 'LPMax'], df_eff.loc[df_eff.index[i], 'eff_index']]
        for j, val in enumerate(values):
            loc_eff_col_char = get_column_letter(col_num + j)
            ws[f'{loc_eff_col_char}{loc_eff_row + i + 1}'] = val

            # Format body
            ws[f'{loc_eff_col_char}{loc_eff_row + i + 1}'].fill = PatternFill(start_color="D9D9D9",
                                                                              end_color="D9D9D9", fill_type="solid")
            ws[f'{loc_eff_col_char}{loc_eff_row + i + 1}'].font = Font(color="0000FF")


def build_report_section2(ws, loc):
    file_in_use = file_preprocessed
    df = pd.read_csv(f'{dir_data_in_use}/{file_in_use}', index_col=0)
    row_num, col_num = loc

    loc_row = row_num
    # --- Title row ---
    headers = ['Units'] + inputName + outputName
    for i, title in enumerate(headers):
        loc_col_char = get_column_letter(col_num + i)
        ws[f'{loc_col_char}{loc_row}'] = title

        # Format header
        ws[f'{loc_col_char}{loc_row}'].fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
        ws[f'{loc_col_char}{loc_row}'].font = Font(color="FFFFFF", bold=True)

    # --- Body rows ---
    for j in range(df.shape[0]):
        values = [df.index[j]] + [df.loc[df.index[j], name] for name in inputName + outputName]
        for i, val in enumerate(values):
            loc_col_char = get_column_letter(col_num + i)
            ws[f'{loc_col_char}{loc_row + j + 1}'] = val

            # Format body
            ws[f'{loc_col_char}{loc_row + j + 1}'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9",
                                                                      fill_type="solid")
            ws[f'{loc_col_char}{loc_row + j + 1}'].font = Font(color="0000FF")

def style_table(ws, start_row, start_col, nrows, ncols):
    """Apply formatting to a table block."""
    # Styles
    header_fill = PatternFill("solid", fgColor="000080")  # Dark blue
    header_font = Font(color="FFFFFF", bold=True)         # White bold
    body_fill = PatternFill("solid", fgColor="D9D9D9")    # Grey
    body_font = Font(color="0000FF")                      # Blue text

    # Header row
    for i in range(ncols):
        col_letter = get_column_letter(start_col + i)
        cell = ws[f"{col_letter}{start_row}"]
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Body rows
    for r in range(1, nrows):
        for i in range(ncols):
            col_letter = get_column_letter(start_col + i)
            cell = ws[f"{col_letter}{start_row + r}"]
            cell.fill = body_fill
            cell.font = body_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

def build_report_section3(ws, loc=[15, 2]):
    file_in_use = file_optimized_LPModel
    df_LPModel = pd.read_csv(f'{dir_data_in_use}/{file_in_use}', index_col=0)

    row_num, col_num = loc

    # --------- Main Title ---------
    ws[f'{get_column_letter(col_num)}{row_num}'] = 'Value from LP Model'
    ws[f'{get_column_letter(col_num)}{row_num}'].font = Font(bold=True)

    # --------- Table Headers ---------
    start_row = row_num + 2
    start_col = col_num

    ws[f'{get_column_letter(start_col)}{start_row}'] = 'Units'

    for i in range(len(inputName)):
        ws[f'{get_column_letter(start_col + i + 1)}{start_row}'] = inputName[i]

    for i in range(len(outputName)):
        ws[f'{get_column_letter(start_col + len(inputName) + i + 1)}{start_row}'] = outputName[i]

    # --------- Table Body ---------
    for j in range(df_LPModel.shape[0]):
        # Units column
        ws[f'{get_column_letter(start_col)}{start_row + j + 1}'] = df_LPModel.index[j]

        # Inputs
        for i in range(len(inputName)):
            val = df_LPModel.loc[df_LPModel.index[j], inputName[i]]
            ws[f'{get_column_letter(start_col + i + 1)}{start_row + j + 1}'] = round(val, 3)

        # Outputs
        for i in range(len(outputName)):
            val = df_LPModel.loc[df_LPModel.index[j], outputName[i]]
            ws[f'{get_column_letter(start_col + len(inputName) + i + 1)}{start_row + j + 1}'] = round(val, 3)

    # --------- Apply Formatting ---------
    nrows = df_LPModel.shape[0] + 1  # header + body
    ncols = 1 + len(inputName) + len(outputName)
    style_table(ws, start_row, start_col, nrows, ncols)

if __name__ == '__main__':
    load_data()
    preprocess_data()
    build_model()
    prepare_report_data()
    generate_report()